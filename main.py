import asyncio
import datetime
import json as _json
from pathlib import Path

import typer
from loguru import logger
from rich.console import Console
from rich.table import Table

from config.settings import settings
from src.database.db_manager import init_db, get_session
from src.database.models import Job, HiringManager

app = typer.Typer(help="Job-Hunt-Forge — AI-powered job search automation.")
console = Console()

_DRIVE = Path("/home/kratosvil/Desarrollo/gdrive/proyectos/JOB-HUNT-FORGE")
_REGION_COLORS = {"USA": "E8F4FD", "Europa": "E8FBE8", "LATAM": "FFF8E8"}


def _recruiter_message(title: str, company: str, matched_skills_json: str) -> str:
    skills = _json.loads(matched_skills_json or "[]")[:3]
    s = " / ".join(skills) if skills else "AI Infrastructure & AWS DevOps"
    return (
        f"Hi! I came across the {title} role at {company}. "
        f"My background in {s} aligns closely — "
        f"I'd love to connect and learn more. — Samir"
    )


def _save_excel_easy_apply(rows: list[dict], path: str, header_color: str = "1F4E79") -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Easy Apply"
    headers = ["#", "Región", "Fit", "Título", "Empresa", "URL", "Estado", "Notas"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = PatternFill("solid", fgColor=header_color)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    for i, w in enumerate([4, 8, 6, 46, 26, 65, 14, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    for i, r in enumerate(rows, 2):
        bg = _REGION_COLORS.get(r["region"], "FFFFFF")
        ws.append([i - 1, r["region"], f"{r['fit']:.0%}", r["title"], r["company"], r["url"], "", ""])
        ws.row_dimensions[i].height = 28
        for col in range(1, 9):
            c = ws.cell(i, col)
            c.border = _border()
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="center" if col in (1, 2, 3, 7) else "left",
                vertical="center", wrap_text=True,
            )
            if col == 6:
                c.hyperlink = r["url"]
                c.font = Font(color="0563C1", underline="single", size=10)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"
    note = ws.cell(len(rows) + 3, 1)
    note.value = f"Generado: {datetime.date.today()} · {len(rows)} jobs · sin repetir aplicados/cerrados"
    note.font = Font(italic=True, color="888888", size=9)
    wb.save(path)


def _save_excel_top_jobs(rows: list[dict], path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Best Fit — Outreach"
    headers = ["#", "Fit", "Título", "Empresa", "URL", "Mensaje Reclutador", "Estado", "Notas"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = PatternFill("solid", fgColor="1A3A52")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    for i, w in enumerate([4, 6, 44, 26, 60, 80, 14, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    alt = "EBF3FB"
    for i, r in enumerate(rows, 2):
        bg = alt if i % 2 == 0 else None
        ws.append([i - 1, f"{r['fit']:.0%}", r["title"], r["company"], r["url"], r["mensaje"], "", ""])
        ws.row_dimensions[i].height = 40
        for col in range(1, 9):
            c = ws.cell(i, col)
            c.border = _border()
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="center" if col in (1, 2, 7) else "left",
                vertical="center", wrap_text=True,
            )
            if col == 5:
                c.hyperlink = r["url"]
                c.font = Font(color="0563C1", underline="single", size=10)
    ws.freeze_panes = "A2"
    note = ws.cell(len(rows) + 3, 1)
    note.value = "Estrategia: abre URL → busca hiring manager → copia el Mensaje Reclutador → Connect con nota (Premium)"
    note.font = Font(italic=True, color="888888", size=9)
    wb.save(path)


@app.callback(invoke_without_command=True)
def startup(ctx: typer.Context) -> None:
    init_db()
    if ctx.invoked_subcommand is None:
        typer.echo(ctx.get_help())


@app.command()
def scrape(
    source: str = typer.Option("linkedin", help="linkedin | indeed | all"),
    limit: int = typer.Option(20, "--limit", help="Max new jobs to analyze per run (0=unlimited)."),
) -> None:
    """Scrape job listings and store them in the database."""
    from src.scrapers.linkedin_scraper import LinkedInScraper
    from src.scrapers.indeed_scraper import IndeedScraper
    from src.intelligence.jd_analyzer import analyze
    import json
    from sqlalchemy.exc import IntegrityError
    from src.database.models import JobStatus

    async def _run():
        scrapers = []
        if source in ("linkedin", "all"):
            scrapers.append(LinkedInScraper())
        if source in ("indeed", "all"):
            scrapers.append(IndeedScraper())

        analyzed = 0

        for scraper in scrapers:
            async with scraper as s:
                async for raw in s.scrape():
                    if limit and analyzed >= limit:
                        logger.info(f"Limit of {limit} jobs reached — stopping scrape.")
                        return
                    if not raw.get("url") or not raw.get("jd_text"):
                        continue
                    try:
                        analysis = analyze(raw["jd_text"])
                        job = Job(
                            url=raw["url"],
                            title=raw["title"].strip(),
                            company=raw["company"].strip(),
                            location=raw.get("location", ""),
                            jd_text=raw["jd_text"],
                            fit_score=analysis.get("fit_score"),
                            matched_skills=json.dumps(analysis.get("matched_skills", [])),
                            missing_skills=json.dumps(analysis.get("missing_skills", [])),
                            growth_signals=json.dumps(analysis.get("company_growth_signals", [])),
                            hiring_manager_signals=json.dumps(analysis.get("hiring_manager_signals", [])),
                            salary_range=analysis.get("salary_range_visible"),
                            remote_friendly=analysis.get("remote_friendly"),
                            source=raw["source"],
                            status=JobStatus.ANALYZED
                            if analysis.get("application_recommended")
                            else JobStatus.SKIPPED,
                        )
                        with get_session() as session:
                            session.add(job)
                        analyzed += 1
                        logger.info(
                            f"[{analyzed}/{limit or '∞'}] Saved: {job.title} @ {job.company} "
                            f"(fit={job.fit_score:.2f}, recommended={analysis.get('application_recommended')})"
                        )
                    except IntegrityError:
                        logger.debug(f"Duplicate skipped: {raw['url']}")
                    except Exception as exc:
                        logger.error(f"Error processing job: {exc}")

    asyncio.run(_run())


@app.command()
def outreach() -> None:
    """Send personalized LinkedIn messages to hiring managers."""
    from src.executors.outreach_bot import OutreachBot

    async def _run():
        async with OutreachBot() as bot:
            sent = await bot.run()
            console.print(f"[green]Outreach complete — {sent} messages sent.[/green]")

    asyncio.run(_run())


@app.command()
def analyze() -> None:
    """Re-analyze all pending jobs in the database."""
    console.print("[yellow]Re-analysis of pending jobs — coming in next phase.[/yellow]")


@app.command()
def apply() -> None:
    """Auto-fill and queue job applications (Phase 2)."""
    console.print("[yellow]Auto-apply — Phase 2 feature. Run outreach first.[/yellow]")


@app.command(name="easy-apply")
def easy_apply(
    usa: int = typer.Option(8, "--usa", help="Max Easy Apply jobs from USA (25%)."),
    europe: int = typer.Option(15, "--europe", help="Max Easy Apply jobs from Europe (50%)."),
    latam: int = typer.Option(7, "--latam", help="Max Easy Apply jobs from LATAM (25%)."),
    min_display_fit: float = typer.Option(0.90, "--min-display-fit", help="Min fit to show in output (DB stores all ≥0.65)."),
    output: str = typer.Option("data/easy_apply.txt", "--output", help="Output TXT file path."),
    excel: str = typer.Option("", "--excel", help="Path to save Excel file (empty = no Excel)."),
    roles: str = typer.Option("", "--roles", help="Comma-separated roles to search (overrides settings)."),
) -> None:
    """Scan LinkedIn Easy Apply jobs (48h) by region — Europa 50% / USA 25% / LATAM 25%."""
    from src.scrapers.easy_apply_scraper import EasyApplyScraper
    from src.intelligence.jd_analyzer import analyze
    from src.database.models import Job, JobStatus
    from sqlalchemy.exc import IntegrityError
    import json

    REGIONS = [
        ("USA",    ["United States"],              usa),
        ("Europa", ["Europe"],                     europe),
        ("LATAM",  ["Colombia", "Latin America"],  latam),
    ]
    custom_roles = [r.strip() for r in roles.split(",") if r.strip()] if roles else None

    async def _scrape_region(locations: list[str], quota: int, region: str) -> list[dict]:
        results = []
        scanned = 0
        scan_limit = quota * 4  # scan up to 4× the quota to find enough recommended

        async with EasyApplyScraper(locations=locations, roles=custom_roles) as scraper:
            async for raw in scraper.scrape():
                if scanned >= scan_limit or len(results) >= quota:
                    break
                if not raw.get("url") or not raw.get("jd_text"):
                    continue

                url = raw["url"]
                scanned += 1

                with get_session() as session:
                    existing = session.query(Job).filter(Job.url == url).first()
                    if existing:
                        # Skip jobs already applied to or marked as closed
                        skip_statuses = (JobStatus.APPLIED, JobStatus.SKIPPED,
                                         JobStatus.REJECTED, JobStatus.OFFER)
                        if existing.status in skip_statuses:
                            continue
                        if existing.source == "linkedin_easy_apply" and existing.fit_score:
                            results.append({
                                "region": region, "url": url,
                                "title": existing.title, "company": existing.company,
                                "fit": existing.fit_score,
                            })
                        continue

                try:
                    analysis = analyze(raw["jd_text"])
                    fit = analysis.get("fit_score", 0.0)
                    recommended = analysis.get("application_recommended", False)

                    job = Job(
                        url=url, title=raw["title"].strip(), company=raw["company"].strip(),
                        location=raw.get("location", ""), jd_text=raw["jd_text"],
                        fit_score=fit,
                        matched_skills=json.dumps(analysis.get("matched_skills", [])),
                        missing_skills=json.dumps(analysis.get("missing_skills", [])),
                        growth_signals=json.dumps(analysis.get("company_growth_signals", [])),
                        hiring_manager_signals=json.dumps(analysis.get("hiring_manager_signals", [])),
                        salary_range=analysis.get("salary_range_visible"),
                        remote_friendly=analysis.get("remote_friendly"),
                        source="linkedin_easy_apply",
                        status=JobStatus.ANALYZED if recommended else JobStatus.SKIPPED,
                    )
                    with get_session() as session:
                        session.add(job)

                    if recommended:
                        results.append({
                            "region": region, "url": url,
                            "title": raw["title"].strip(), "company": raw["company"].strip(),
                            "fit": fit,
                        })
                    logger.info(f"[{region}][{scanned}] {raw['title'][:40]} @ {raw['company'][:20]} fit={fit:.0%} {'✓' if recommended else '✗'}")
                except IntegrityError:
                    logger.debug(f"Duplicate skipped: {url}")
                except Exception as exc:
                    logger.error(f"Error processing {url}: {exc}")

        results.sort(key=lambda x: x["fit"], reverse=True)
        return results[:quota]

    async def _run():
        all_results = []
        total_scanned = 0

        for region_name, locs, quota in REGIONS:
            logger.info(f"--- Scanning {region_name} (quota: {quota}) ---")
            region_results = await _scrape_region(locs, quota, region_name)
            all_results.extend(region_results)
            total_scanned += len(region_results)
            logger.info(f"--- {region_name}: {len(region_results)}/{quota} found ---")

        # Filter for display — only ≥ min_display_fit
        display = [r for r in all_results if r["fit"] >= min_display_fit]
        display.sort(key=lambda x: ({"USA": 2, "Europa": 0, "LATAM": 1}.get(x["region"], 9), -x["fit"]))

        # Terminal table
        table = Table(title=f"Easy Apply ≥{min_display_fit:.0%} — {len(display)} jobs · 48h · {datetime.date.today()}")
        table.add_column("#",      style="dim",     width=3)
        table.add_column("Región", style="yellow",  width=7)
        table.add_column("Fit",    style="green",   width=6)
        table.add_column("Título", style="cyan",    min_width=35)
        table.add_column("Empresa",style="magenta", min_width=18)
        table.add_column("URL")

        for i, r in enumerate(display, 1):
            table.add_row(str(i), r["region"], f"{r['fit']:.0%}",
                          r["title"][:48], r["company"][:25], r["url"])
        console.print(table)

        # TXT output
        with open(output, "w") as f:
            f.write(f"Easy Apply ≥{min_display_fit:.0%} — {len(display)} jobs · 48h · {datetime.date.today()}\n")
            f.write("=" * 70 + "\n\n")
            for i, r in enumerate(display, 1):
                f.write(f"{i}. [{r['region']}][{r['fit']:.0%}] {r['title']} @ {r['company']}\n")
                f.write(f"   {r['url']}\n\n")
        console.print(f"[green]Saved {len(display)} links (≥{min_display_fit:.0%}) → {output}[/green]")

        # Excel output
        if excel:
            _save_excel_easy_apply(display, excel)
            console.print(f"[green]Excel → {excel}[/green]")

        return display

    results = asyncio.run(_run())
    return results


@app.command(name="top-jobs")
def top_jobs(
    top: int = typer.Option(10, "--top", help="Number of best-fit jobs to show."),
    min_fit: float = typer.Option(0.80, "--min-fit", help="Minimum fit score threshold."),
    output: str = typer.Option("data/top_jobs.txt", "--output", help="Output TXT file path."),
    excel: str = typer.Option("", "--excel", help="Path to save Excel file with recruiter messages."),
) -> None:
    """Top N best-fit non-Easy-Apply jobs from DB — includes recruiter message for outreach."""
    from src.database.models import Job, JobStatus

    with get_session() as session:
        jobs = (
            session.query(Job)
            .filter(
                Job.source != "linkedin_easy_apply",
                Job.fit_score >= min_fit,
                Job.status == JobStatus.ANALYZED,
            )
            .order_by(Job.fit_score.desc())
            .limit(top)
            .all()
        )

        if not jobs and min_fit > 0.65:
            jobs = (
                session.query(Job)
                .filter(
                    Job.source != "linkedin_easy_apply",
                    Job.fit_score >= settings.min_fit_score,
                    Job.status == JobStatus.ANALYZED,
                )
                .order_by(Job.fit_score.desc())
                .limit(top)
                .all()
            )

        results = [
            {
                "title": j.title, "company": j.company, "fit": j.fit_score,
                "location": j.location or "", "url": j.url,
                "mensaje": _recruiter_message(j.title, j.company, j.matched_skills or "[]"),
            }
            for j in jobs
        ]

    table = Table(title=f"Best Fit (no Easy Apply) — Top {len(results)} · fit ≥ {min_fit:.0%}")
    table.add_column("#",       style="dim",     width=3)
    table.add_column("Fit",     style="green",   width=6)
    table.add_column("Título",  style="cyan",    min_width=35)
    table.add_column("Empresa", style="magenta", min_width=20)
    table.add_column("URL")

    for i, r in enumerate(results, 1):
        table.add_row(str(i), f"{r['fit']:.0%}", r["title"][:45],
                      r["company"][:25], r["url"])
    console.print(table)

    with open(output, "w") as f:
        f.write(f"Best Fit Jobs (manual outreach) — Top {len(results)} · {datetime.date.today()}\n")
        f.write("=" * 70 + "\n\n")
        for i, r in enumerate(results, 1):
            f.write(f"{i}. [{r['fit']:.0%}] {r['title']} @ {r['company']}\n")
            f.write(f"   URL:     {r['url']}\n")
            f.write(f"   Mensaje: {r['mensaje']}\n\n")

    console.print(f"[green]Saved {len(results)} links → {output}[/green]")

    if excel:
        _save_excel_top_jobs(results, excel)
        console.print(f"[green]Excel → {excel}[/green]")


@app.command(name="find-recruiters")
def find_recruiters(
    max_per_query: int = typer.Option(8, "--max-per-query", help="Max recruiters to extract per search query."),
    output: str = typer.Option("data/recruiters.txt", "--output", help="Output TXT file path."),
    excel: str = typer.Option("", "--excel", help="Path to save Excel file in Drive."),
) -> None:
    """Search LinkedIn People for tech recruiters (DevOps/MLOps/Cloud) and generate outreach messages."""
    from src.scrapers.recruiter_finder_scraper import RecruiterFinderScraper

    def _build_message(name: str, company: str) -> str:
        first = name.split()[0] if name else "Hola"
        comp = f"en {company}" if company and len(company) < 40 else "en tu empresa"
        return (
            f"Hola {first}, soy Senior DevOps / AI Infrastructure Engineer con experiencia "
            f"en AWS, Bedrock y MLOps. Si {comp} tienen roles remotos en este perfil, "
            f"me encantaría estar en tu red. — Samir"
        )

    async def _run() -> list[dict]:
        seen_urls = set()
        results = []

        async with RecruiterFinderScraper(max_per_query=max_per_query) as scraper:
            async for person in scraper.scrape():
                url = person.get("linkedin_url", "")
                if not url or url in seen_urls:
                    continue
                seen_urls.add(url)
                person["mensaje"] = _build_message(person["name"], person["company"])
                results.append(person)
                logger.info(
                    f"  + {person['name']} | {person['title'][:40]} @ {person['company'][:25]}"
                )

        return results

    results = asyncio.run(_run())

    # Terminal table
    table = Table(title=f"Recruiters encontrados — {len(results)} · {datetime.date.today()}")
    table.add_column("#",        style="dim",     width=3)
    table.add_column("Nombre",   style="cyan",    min_width=22)
    table.add_column("Empresa",  style="magenta", min_width=20)
    table.add_column("Rol",      style="yellow",  min_width=28)
    table.add_column("URL")
    for i, r in enumerate(results, 1):
        table.add_row(str(i), r["name"][:25], r["company"][:22],
                      r["title"][:30], r["linkedin_url"])
    console.print(table)

    # TXT
    with open(output, "w") as f:
        f.write(f"Recruiters Tech (DevOps/MLOps/Cloud) — {len(results)} · {datetime.date.today()}\n")
        f.write("=" * 70 + "\n\n")
        for i, r in enumerate(results, 1):
            f.write(f"{i}. {r['name']} | {r['title']} @ {r['company']}\n")
            f.write(f"   URL:     {r['linkedin_url']}\n")
            f.write(f"   Mensaje: {r['mensaje']}\n\n")
    console.print(f"[green]Saved {len(results)} recruiters → {output}[/green]")

    # Excel
    if excel and results:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        def _border():
            s = Side(style="thin", color="BFBFBF")
            return Border(left=s, right=s, top=s, bottom=s)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recruiters"
        headers = ["#", "Nombre", "Empresa", "Rol en LinkedIn", "URL Perfil",
                   "Mensaje (copiar como nota)", "Enviado", "Respuesta", "Notas"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(1, col)
            c.fill = PatternFill("solid", fgColor="2E4057")
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
        for i, w in enumerate([4, 24, 24, 34, 55, 85, 10, 12, 24], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 20
        for i, r in enumerate(results, 2):
            ws.append([
                i - 1, r["name"], r["company"], r["title"],
                r["linkedin_url"], r["mensaje"], "☐", "", "",
            ])
            ws.row_dimensions[i].height = 45
            for col in range(1, 10):
                c = ws.cell(i, col)
                c.border = _border()
                bg = "F0F4F8" if i % 2 == 0 else None
                if bg:
                    c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(
                    horizontal="center" if col in (1, 7) else "left",
                    vertical="center", wrap_text=True,
                )
                if col == 5:
                    c.hyperlink = r["linkedin_url"]
                    c.font = Font(color="0563C1", underline="single", size=10)
        ws.freeze_panes = "A2"
        note = ws.cell(len(results) + 3, 1)
        note.value = (
            "Instrucciones: abre cada URL → Connect → pega el Mensaje como nota (Premium). "
            f"Generado: {datetime.date.today()} · {len(results)} reclutadores únicos."
        )
        note.font = Font(italic=True, color="888888", size=9)
        wb.save(excel)
        console.print(f"[green]Excel → {excel}[/green]")


@app.command()
def pipeline(
    skip_scrape: bool = typer.Option(False, "--skip-scrape", help="Skip scraping stage."),
    skip_outreach: bool = typer.Option(False, "--skip-outreach", help="Skip outreach stage."),
    source: str = typer.Option("linkedin", help="linkedin | indeed | all"),
    limit: int = typer.Option(0, "--limit", help="Max new jobs to analyze (0=unlimited). Use for test runs."),
    managers_limit: int = typer.Option(0, "--managers-limit", help="Max jobs to search managers for per run (0=unlimited)."),
) -> None:
    """Run the full pipeline: scrape → find managers → outreach → report."""
    from src.pipeline import run as run_pipeline
    asyncio.run(run_pipeline(
        skip_scrape=skip_scrape,
        skip_outreach=skip_outreach,
        source=source,
        limit=limit,
        managers_limit=managers_limit,
    ))


@app.command()
def status() -> None:
    """Show current database statistics."""
    with get_session() as session:
        jobs = session.query(Job).all()
        managers = session.query(HiringManager).all()

    table = Table(title="Job-Hunt-Forge Status")
    table.add_column("Metric", style="cyan")
    table.add_column("Count", style="magenta")

    from collections import Counter
    job_counts = Counter(j.status.value for j in jobs)
    for status_name, count in sorted(job_counts.items()):
        table.add_row(f"Jobs — {status_name}", str(count))

    mgr_counts = Counter(m.status.value for m in managers)
    for status_name, count in sorted(mgr_counts.items()):
        table.add_row(f"Managers — {status_name}", str(count))

    table.add_row("Total jobs", str(len(jobs)))
    table.add_row("Total managers", str(len(managers)))
    console.print(table)


if __name__ == "__main__":
    app()
